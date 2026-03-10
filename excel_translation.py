import openpyxl
import os
from openpyxl.utils import range_boundaries
from deep_translator import GoogleTranslator
import concurrent.futures
import threading
import time

# ---------- 配置 ----------
DEBUG = True          # True 时打印调试信息
MAX_RETRIES = 2       # 翻译失败重试次数
RETRY_DELAY = 0.5     # 重试间隔（秒）
# --------------------------

# 词典
test_dictionary = {
    # 组合词（需要在单个词之前处理）
    "新增成功": "Successfully Added",
    "导入新增": "Import and Add",
    "导入成功": "Import Successful",
    "保存成功": "Save Successful",
    "设备成功": "Equipment Successfully",
    "新增or编辑": "Add or Edit",
    "设备类型分类": "Equipment Type Classification",
    "设备台账": "Equipment Ledger",
    # 单个词
    "保存": "Save",
    "删除": "Delete",
    "确认": "Confirm",
    "取消": "Cancel",
    "新增": "Add",
    "编辑": "Edit",
    "查询": "Query",
    "导出": "Export",
    "导入": "Import",
    "设置": "Settings",
    "未开始": "Not Started",
    "进行中": "In Progress",
    "已完成": "Completed",
    "通过": "Passed",
    "失败": "Failed",
    "阻塞": "Blocked",
    "跳过": "Skipped",
    "成功": "Success",
    "运营平台": "Operations Platform",
    "点位管理": "IoT Management",
    "物联网管理": "IoT Management",
    "设备厂商": "Device Manufacturer",
    "数据采集网关": "Data Gateway",
    "传感器": "Sensor",
    "点位来源": "Data Point Source",
    "普通表计": "Traditional Meter",
    "智能表计": "Smart Meter",
    "客户": "Client",
    "站点": "Site",
    "设备": "Equipment",
    "设备类型": "Equipment Type",
    "测试用例": "Test Case",
    "前置条件": "Preconditions",
    "测试步骤": "Test Steps",
    "预期结果": "Expected Results",
    "实际结果": "Actual Results",
    "测试环境": "Test Environment",
    "功能测试": "Functional Testing",
    "回归测试": "Regression Testing",
    "冒烟测试": "Smoke Testing",
    "性能测试": "Performance Testing",
}

translation_cache = {}
cache_lock = threading.Lock()
_thread_local = threading.local()

def _get_thread_translator():
    if not hasattr(_thread_local, "translator"):
        _thread_local.translator = GoogleTranslator(source='zh-CN', target='en')
    return _thread_local.translator

def normalize_key(text: str) -> str:
    return text.strip()

def dict_replace_whole_sentence(text: str) -> str:
    """使用词典替换，使用占位符避免重复替换"""
    import re
    replaced = text
    replacements = []
    
    # 按长度从长到短排序，避免短词先替换导致长词无法匹配
    for idx, (zh_term, en_term) in enumerate(sorted(test_dictionary.items(), key=lambda kv: -len(kv[0]))):
        # 使用带编号的占位符，确保唯一性
        placeholder = f"__PLACEHOLDER_{idx}__"
        if zh_term in replaced:
            replaced = replaced.replace(zh_term, placeholder)
            replacements.append((placeholder, en_term))
    
    # 用实际的英文替换占位符
    for placeholder, en_term in replacements:
        replaced = replaced.replace(placeholder, en_term)
    
    return replaced

def translate_with_retry(text: str):
    last_exc = None
    for attempt in range(MAX_RETRIES + 1):
        try:
            translator = _get_thread_translator()
            return translator.translate(text)
        except Exception as e:
            last_exc = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY * (attempt + 1))
            else:
                raise
    raise last_exc

def process_text(text):
    key = normalize_key(text)
    with cache_lock:
        if key in translation_cache:
            if DEBUG:
                print(f"[CACHE HIT] {text!r} -> {translation_cache[key]!r}")
            return translation_cache[key]

    replaced = dict_replace_whole_sentence(text)
    if any('\u4e00' <= c <= '\u9fff' for c in replaced):
        try:
            translated = translate_with_retry(replaced)
        except Exception as e:
            translated = f"[Translation Failed]: {str(e)}"
    else:
        translated = replaced

    # Use lock to ensure debug output is printed atomically
    if DEBUG:
        with cache_lock:
            print(f"[DEBUG] 原文: {text!r}")
            print(f"[DEBUG] 替换后: {replaced!r}")
            print(f"[DEBUG] 最终翻译: {translated!r}")
            print("-----")

    with cache_lock:
        translation_cache[key] = translated
    return translated

def translate_excel(input_file, output_file, columns_to_translate=None, skip_header=True, max_workers=8, dedupe=True):
    print(f"加载Excel文件: {input_file}")
    # 保留所有格式和样式
    wb = openpyxl.load_workbook(input_file, data_only=False, keep_vba=False)
    ws = wb.active

    # 1️⃣ 先解除所有合并单元格，并记录主单元格的值
    print(f"原始文件: 行={ws.max_row}, 列={ws.max_column}, 合并单元格={len(list(ws.merged_cells))}")
    merged_ranges = list(ws.merged_cells.ranges)
    merge_map = []  # 存储 (min_row, min_col, max_row, max_col)
    
    # 先记录所有合并信息和主单元格的值
    for mc_range in merged_ranges:
        min_row, min_col, max_row, max_col = mc_range.min_row, mc_range.min_col, mc_range.max_row, mc_range.max_col
        master_value = ws.cell(row=min_row, column=min_col).value
        merge_map.append((min_row, min_col, max_row, max_col, master_value))
    
    # 再解除合并
    for mc_range in merged_ranges:
        ws.unmerge_cells(str(mc_range))
    
    print(f"已记录 {len(merge_map)} 个合并单元格信息")

    # 2️⃣ 收集所有含中文的单元格
    start_row = 2 if skip_header else 1
    to_translate = []
    cell_coords = []

    for row_idx in range(start_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            if columns_to_translate and col_idx not in columns_to_translate:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if not cell.value or not isinstance(cell.value, str):
                continue
            if not any('\u4e00' <= c <= '\u9fff' for c in cell.value):
                continue
            if DEBUG:
                print(f"收集({row_idx},{col_idx}): {cell.value}")
            to_translate.append(cell.value)
            cell_coords.append((row_idx, col_idx))

    print(f"开始翻译，共 {len(to_translate)} 项，使用 {max_workers} 线程")

    if dedupe:
        unique_texts = list(dict.fromkeys(to_translate))
        translation_map = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(process_text, txt): txt for txt in unique_texts}
            for i, fut in enumerate(concurrent.futures.as_completed(futures), 1):
                src = futures[fut]
                try:
                    res = fut.result()
                except Exception as e:
                    res = f"[Translation Failed]: {e}"
                translation_map[src] = res
                # Print progress after DEBUG output to avoid interruption
                if not DEBUG:
                    print(f"已翻译去重项 {i}/{len(unique_texts)}")
        
        # Print final progress summary when DEBUG is enabled
        if DEBUG:
            print(f"\n==> 已完成所有 {len(unique_texts)} 项去重翻译\n")
        
        translated_texts = [translation_map[t] for t in to_translate]
    else:
        translated_texts = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = executor.map(process_text, to_translate)
            for i, result in enumerate(results):
                translated_texts.append(result)
                print(f"进度: {i + 1}/{len(to_translate)}")

    # 3️⃣ 回填翻译结果
    for coord, translation in zip(cell_coords, translated_texts):
        row_idx, col_idx = coord
        ws.cell(row=row_idx, column=col_idx).value = translation

    # 4️⃣ 重新合并单元格
    print("重新合并单元格...")
    for merge_info in merge_map:
        min_row, min_col, max_row, max_col, original_value = merge_info
        # 获取当前主单元格的值（可能已被翻译）
        current_value = ws.cell(row=min_row, column=min_col).value
        
        # 如果主单元格的值是None，使用原始值
        if current_value is None and original_value is not None:
            ws.cell(row=min_row, column=min_col).value = original_value
        
        # 合并单元格
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    
    print(f"保存前检查: 行={ws.max_row}, 列={ws.max_column}, 合并单元格={len(list(ws.merged_cells))}")

    wb.save(output_file)
    print(f"保存翻译文件: {output_file}")
    print(f"翻译完成，共翻译 {len(to_translate)} 项")

if __name__ == "__main__":
    input_file = 'Metersphere_case_Akila (42).xlsx'
    base, ext = os.path.splitext(input_file)
    output_file = f"{base}_update{ext}"
    columns_to_translate = [2, 4, 6,7]  # 可以设为 None 翻译整张表
    translate_excel(input_file, output_file, columns_to_translate, skip_header=True, max_workers=8, dedupe=True)
